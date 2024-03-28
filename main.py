from flask import Flask, render_template, request, Response
from pymongo import MongoClient
import pandas as pd
import io, os
import openpyxl

def crear_app():
    app = Flask(__name__)
    client = MongoClient(os.getenv("URL_DB"))
    db = client['inscriptos']  # Nombre de tu base de datos
    participantes_collection = db['participantes']  # Colección donde se almacenarán los datos

    @app.route("/", methods=['GET'])
    def show_form():
        print("Mostrando formulario...")
        return render_template("formulario.html")

    @app.route("/descargar_archivo", methods=['GET'])
    def get_inscriptos():
        try:
            print("Descargando archivo...")
            inscriptos = participantes_collection.find()

            output = io.BytesIO()
            writer = pd.ExcelWriter(output, engine='xlsxwriter')
            df = pd.DataFrame(list(inscriptos))
            df.drop('_id', axis=1, inplace=True)  # Eliminar el ID de MongoDB
            df.to_excel(writer, index=False)
            writer.close()
            output.seek(0)

            print("Archivo descargado correctamente.")
            return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename="inscriptos.xlsx"'})

        except Exception as e:
            print(f"Error al descargar el archivo: {e}")
            return str(e), 500

    @app.route("/base", methods=['GET'])
    def show_inscriptos():
        try:
            print("Mostrando inscriptos...")
            inscriptos = participantes_collection.find()

            df = pd.DataFrame(list(inscriptos))
            df.drop('_id', axis=1, inplace=True)  # Eliminar el ID de MongoDB
            
            html_content = df.to_html(index=False)
            
            print("Inscriptos mostrados correctamente.")
            return render_template("mostrar.html", content=html_content)

        except Exception as e:
            print(f"Error al mostrar inscriptos: {e}")
            return str(e), 500

    
    @app.route("/buscar_habitacion", methods=['POST'])
    def buscar_habitacion():
        try:
            print("Buscando habitación...")
            correo = request.form.get('correo')
            inscripto = participantes_collection.find_one({"correo": correo.lower()})

            # Cargar el segundo archivo Excel que contiene los enlaces de las compañías
            libro_excel_enlaces = openpyxl.load_workbook('./archivos/datos_fijos.xlsx')
            hoja_excel_enlaces = libro_excel_enlaces.active

            # Inicializar linkcompania con None por defecto
            linkcompania = None
            color_text = None
            ubic = None

            if inscripto:
                participantes_collection.update_one({"_id": inscripto["_id"]}, {"$set": {"registro": "SI"}})

                # Buscar el enlace correspondiente a la compañía en el segundo archivo Excel
                for fila in hoja_excel_enlaces.iter_rows(min_row=2, values_only=True):
                    if fila[0] == inscripto["compania"] and fila[1] == inscripto["sesion"]:
                        linkcompania = fila[2]
                        break

                # Buscar el color
                for fila in hoja_excel_enlaces.iter_rows(min_row=2, values_only=True):
                    if fila[4] == inscripto["sesion"]:
                        color_text = fila[5]
                        break

                # Buscar la ubicacion de la habitacion
                for fila in hoja_excel_enlaces.iter_rows(min_row=2, values_only=True):
                    if fila[7] == inscripto["habitacion"]:
                        ubic = fila[8]
                        break
                
                color = "red" if inscripto["sesion"] == 1 else "blue" if inscripto["sesion"] == 2 else "yellow"

                print("Habitación encontrada.")
                return render_template("respuesta.html", participante=inscripto["participante"], habitacion=inscripto["habitacion"], compania=inscripto["compania"], consejero=inscripto["consejero"], consejera=inscripto["consejera"], sesion=color, link_compania=linkcompania, color=color_text.upper(), ubicacion=ubic)

            else:
                print("Correo no encontrado.")
                return render_template("error.html"), 404

        except Exception as e:
            print(f"Error al buscar habitación: {e}")
            return str(e), 500
        
    return app

if __name__ == "__main__":
    print("Creando aplicación...")
    app = crear_app()
    print("Aplicación creada.")
    print("Iniciando servidor...")
    app.run()